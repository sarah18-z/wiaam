import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk, ImageDraw
from openpyxl import load_workbook
from datetime import datetime
import copy
import itertools
import random # Added for the random swaps in improvement heuristic
from genetique import GeneticAlgorithmORs, evaluer_sequence

# Import matplotlib components
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches # For legend patches in Gantt chart
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk # For embedding Gantt chart in Tkinter
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


class Application:
    def __init__(self, root):
        self.root = root
        self.root.title("Interface d'Ordonnancement")
        self.root.attributes('-fullscreen', True)
        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()
       
        self.images = {}
        self.saisies_ors = []
        self.donnees_importees = []
       
        self.canvas = tk.Canvas(self.root, width=self.screen_width, height=self.screen_height, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
       
        # Initialize machines_state for the heuristic
        self.machines_state = {
            "Usinage": [{"nom": "M1-Usinage", "libre": 0}, {"nom": "M2-Usinage", "libre": 0}],
            "Culasse": [{"nom": "M1-Culasse", "libre": 0}, {"nom": "M2-Culasse", "libre": 0}],
            "Injection": [{"nom": "M1-Inj", "libre": 0}, {"nom": "M2-Inj", "libre": 0}],
            "BEM": [{"nom": "M1-BEM", "libre": 0}, {"nom": "M2-BEM", "libre": 0}],
            "Peinture": [{"nom": "M1-Peinture", "libre": 0}, {"nom": "M2-Peinture", "libre": 0}]
        }
       
        self.page1()

    def quitter(self):
        self.root.destroy()

   
   

    def page1(self):
        self.canvas.delete("all")
       
        try:
            bg_image = Image.open("ordonn.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg1'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg1'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement image: {e}")
            self.canvas.configure(bg="lightgray")

        frame_width, frame_height = 600, 500
        x0 = 80
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.7)))
            self.images['overlay1'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay1'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        try:
            usthb_img = Image.open("logo usthb.png").resize((80, 80), Image.LANCZOS)
            self.images['usthb_logo'] = ImageTk.PhotoImage(usthb_img)
            self.canvas.create_image(x0 + frame_width - 100, y0 + 10, image=self.images['usthb_logo'], anchor="nw")

            cosider_img = Image.open("logo cosider 2.png").resize((80, 80), Image.LANCZOS)
            self.images['cosider_logo'] = ImageTk.PhotoImage(cosider_img)
            self.canvas.create_image(x0 + 20, y0 + 10, image=self.images['cosider_logo'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement logos: {e}")
            self.canvas.create_text(x0 + 20, y0 + 50, text="COSIDER", font=("Arial", 12), anchor="nw")
            self.canvas.create_text(x0 + frame_width - 100, y0 + 50, text="USTHB", font=("Arial", 12), anchor="nw")

        self.canvas.create_text(x0 + frame_width // 2, y0 + 80, text="Bienvenue", font=("Georgia", 36, "bold italic"), fill="black")

        label_font = ("Arial", 12)
        field_x = x0 + 120
        value_x = x0 + 380

        self.canvas.create_text(field_x, y0 + 180, text="Nom:", font=label_font, fill="black", anchor="w")
        self.nom_entry = ttk.Entry(self.root, width=30)
        self.canvas.create_window(value_x, y0 + 180, window=self.nom_entry)

        self.canvas.create_text(field_x, y0 + 230, text="Prénom:", font=label_font, fill="black", anchor="w")
        self.prenom_entry = ttk.Entry(self.root, width=30)
        self.canvas.create_window(value_x, y0 + 230, window=self.prenom_entry)

        self.canvas.create_text(field_x, y0 + 280, text="Mot de passe:", font=label_font, fill="black", anchor="w")
        self.mdp_entry = ttk.Entry(self.root, width=30, show="*")
        self.canvas.create_window(value_x, y0 + 280, window=self.mdp_entry)

        enter_btn = tk.Button(self.root, text="Entrer", command=self.aller_a_page2,
                               bg="#007BFF", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x0 + 180, y0 + 370, window=enter_btn, width=120)

        quit_btn = tk.Button(self.root, text="Quitter", command=self.quitter,
                               bg="#DC3545", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x0 + 390, y0 + 370, window=quit_btn, width=120)

    def aller_a_page2(self):
        if not self.nom_entry.get() or not self.prenom_entry.get() or not self.mdp_entry.get():
            messagebox.showwarning("Champs manquants", "Veuillez remplir tous les champs")
            return
           
        self.canvas.delete("all")
        self.page2()

    def page2(self):
        self.canvas.delete("all")
       
        try:
            bg_image = Image.open("bg2.webp").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg2'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg2'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg2: {e}")
            self.canvas.configure(bg="lightblue")

        frame_width = 800
        frame_height = 400
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.6)))
            self.images['overlay2'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay2'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        self.canvas.create_text(self.screen_width // 2, y0 + 50,
                                 text="Choisissez votre mode de saisie de donnée",
                                 font=("Georgia", 24, "bold italic"))

        btn1 = tk.Button(self.root, text="Saisie manuelle", font=("Arial", 14, "italic"),
                         command=self.page3, bg="#28a745", fg="white", width=20)
        self.canvas.create_window(self.screen_width // 2, y0 + 120, window=btn1)

        btn2 = tk.Button(self.root, text="Saisie par Excel", font=("Arial", 14, "italic"),
                         command=self.importer_excel, bg="#17a2b8", fg="white", width=20)
        self.canvas.create_window(self.screen_width // 2, y0 + 180, window=btn2)

        quit_btn = tk.Button(self.root, text="Quitter", command=self.quitter,
                               bg="#dc3545", fg="white", font=("Arial", 10, "italic"))
        self.canvas.create_window(self.screen_width // 2, y0 + 300, window=quit_btn, width=120)

    def page3(self):
        self.canvas.delete("all")
       
        try:
            bg_image = Image.open("image_ordo_blog.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg3'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg3'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg3: {e}")
            self.canvas.configure(bg="lightblue")

        frame_width = 1200
        frame_height = 700
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.7)))
            self.images['overlay3'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay3'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        self.canvas.create_text(self.screen_width // 2, y0 + 40,
                                 text="Saisie manuelle des ORs",
                                 font=("Georgia", 24, "bold italic"))

        self.canvas.create_text(x0 + 50, y0 + 90,
                                 text="Nombre d'ORs à saisir:",
                                 font=("Arial", 12), anchor="w")
       
        self.nb_ors_entry = ttk.Entry(self.root, width=10, font=("Arial", 12))
        self.canvas.create_window(x0 + 280, y0 + 90, window=self.nb_ors_entry)

        def valider_nb_ors():
            try:
                nb_ors = int(self.nb_ors_entry.get())
                if nb_ors <= 0:
                    messagebox.showerror("Erreur", "Le nombre d'ORs doit être supérieur à 0")
                else:
                    self.generer_saisie_ors(nb_ors, x0 + 50, y0 + 130)
            except ValueError:
                messagebox.showerror("Erreur", "Veuillez entrer un nombre valide d'ORs")

        bouton_valider = tk.Button(self.root, text="Valider", command=valider_nb_ors,
                                     bg="#007BFF", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x0 + 380, y0 + 90, window=bouton_valider)

        quit_btn = tk.Button(self.root, text="Quitter", command=self.quitter,
                               bg="#DC3545", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x0 + frame_width - 80, y0 + frame_height - 40, window=quit_btn)

    def generer_saisie_ors(self, nb_ors, x_start, y_start):
        self.saisies_ors = []
       
        # Clear previous entry widgets and labels related to OR input
        # This is crucial to avoid multiple widgets piling up on refresh
        for widget in self.canvas.winfo_children():
            # Check if the widget is part of the dynamic OR input area
            # This is a bit of a heuristic check, adjust if needed
            if isinstance(widget, (ttk.Entry, tk.Button)) and widget not in [self.nb_ors_entry]:
                widget.destroy()

        # Remove previous text items related to tasks
        for item in self.canvas.find_all():
            if self.canvas.type(item) == "text" and "DURÉES DES TÂCHES" in self.canvas.itemcget(item, "text"):
                self.canvas.delete(item)
            # Add other text items to delete if they persist from previous calls
            if self.canvas.type(item) == "text" and "OR" in self.canvas.itemcget(item, "text") and "Date limite" not in self.canvas.itemcget(item, "text"):
                 self.canvas.delete(item)


        self.canvas.create_text(x_start + 500, y_start + 10,
                                 text="DURÉES DES TÂCHES ",
                                 font=("Arial", 14, "bold"), anchor="center")
       
        taches_tkinter = ["Demontage", "Usinage", "Culasse", "Injection", "Essai",
                          "Peinture", "Confection", "Montage", "Controle"] # Corrected to match 9 tasks and their names

        espacement_taches = 100
       
        for j, nom_tache in enumerate(taches_tkinter):
            self.canvas.create_text(x_start + 300 + j * espacement_taches, y_start + 40,
                                     text=nom_tache,
                                     font=("Arial", 10, "bold"), anchor="center")

        self.canvas.create_text(x_start + 30, y_start + 80,
                                 text="OR",
                                 font=("Arial", 12, "bold"), anchor="w")
       
        self.canvas.create_text(x_start + 100, y_start + 80,
                                 text="Date limite\n(jj/mm/aaaa)",
                                 font=("Arial", 12, "bold"), anchor="w")

        for i in range(nb_ors):
            y = y_start + 110 + i * 40
           
            self.canvas.create_text(x_start + 40, y,
                                     text=f"OR {i + 1}",
                                     font=("Arial", 11), anchor="w")
           
            date_entry = ttk.Entry(self.root, width=12, font=("Arial", 10))
            self.canvas.create_window(x_start + 150, y, window=date_entry)
           
            durees = []
            for j in range(len(taches_tkinter)):
                entry = ttk.Entry(self.root, width=5, font=("Arial", 10))
                self.canvas.create_window(x_start + 300 + j * espacement_taches, y, window=entry)
                durees.append(entry)
           
            self.saisies_ors.append({"date_limite": date_entry, "durees": durees})
       
        y_boutons = y_start + 110 + nb_ors * 40 + 30
       
        bouton_retour = tk.Button(self.root, text="Retour", command=self.page2,
                                     bg="#6C757D", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x_start + 800, y_boutons, window=bouton_retour, width=100)
       
        bouton_suivant = tk.Button(self.root, text="Suivant",
                                     command=lambda: self.page4(self.recuperer_donnees_saisie()),
                                     bg="#28A745", fg="white", font=("Arial", 10, "bold"))
        self.canvas.create_window(x_start + 900, y_boutons, window=bouton_suivant, width=100)

    def recuperer_donnees_saisie(self):
        donnees = []
        for i, or_data in enumerate(self.saisies_ors):
            try:
                durees = [float(entry.get()) if entry.get() else 0.0 for entry in or_data["durees"]]
                donnees.append({
                    "nom": f"OR{i + 1}",
                    "date_limite": or_data["date_limite"].get(),
                    "durees": durees
                })
            except ValueError:
                messagebox.showerror("Erreur", "Veuillez entrer des valeurs numériques pour les durées")
                return None
        return donnees

    def importer_excel(self):
        fichier_excel = filedialog.askopenfilename(
            title="Choisir un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls")]
        )
       
        if not fichier_excel:
            return
       
        try:
            wb = load_workbook(filename=fichier_excel)
            ws = wb.active
            self.donnees_importees = []
           
            # Assuming first row is header, data starts from second row
            # Columns: OR Name, Due Date, Task1_Duration, ..., Task9_Duration
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: # Stop if OR name is empty
                    break
                   
                or_name = str(row[0]) # e.g., "OR1"
               
                date_limite = row[1]
                if isinstance(date_limite, datetime):
                    date_limite = date_limite.strftime("%d/%m/%Y")
                elif not isinstance(date_limite, str):
                    date_limite = str(date_limite)
               
                durees = []
                # Assuming 9 duration columns from column 3 (index 2) to column 11 (index 10)
                for duree_val in row[2:11]:  # Slices up to but not including 11
                    try:
                        durees.append(float(duree_val) if duree_val is not None else 0.0)
                    except (ValueError, TypeError):
                        durees.append(0.0)
               
                or_data = {
                    "nom": or_name,
                    "date_limite": date_limite,
                    "durees": durees
                }
                self.donnees_importees.append(or_data)
           
            wb.close()
           
            if not self.donnees_importees:
                messagebox.showwarning("Avertissement", "Aucune donnée valide trouvée dans le fichier.")
                return
           
            messagebox.showinfo("Succès", f"{len(self.donnees_importees)} ORs importés avec succès!")
            self.page4(self.donnees_importees)
           
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture du fichier:\n{str(e)}")


    def page4(self, donnees=None):
        self.canvas.delete("all")
       
        try:
            bg_image = Image.open("ordonnancement.png").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg4'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg4'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg4: {e}")
            self.canvas.configure(bg="lightblue")

        frame_width = 1200
        frame_height = 700
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.7)))
            self.images['overlay4'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay4'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        self.canvas.create_text(self.screen_width // 2, y0 + 40,
                                 text="RÉSULTATS DE L'ORDONNANCEMENT",
                                 font=("Georgia", 24, "bold italic"), fill="black")

        data = donnees if donnees is not None else self.donnees_importees
       
        if not data:
            self.canvas.create_text(self.screen_width // 2, y0 + 100,
                                     text="Aucune donnée à afficher",
                                     font=("Arial", 16), fill="red")
            return

        self.canvas.create_text(x0 + 50, y0 + 80,
                                 text=f"{len(data)} ORs à traiter:",
                                 font=("Arial", 14, "bold"), anchor="w")

        # Display tasks as aligned with the improvement code's OR structure
        taches_display = ["Demontage", "Usinage", "Culasse", "Injection", "Essai",
                          "Peinture", "Confection", "Montage", "Controle"]

        or_colors_rgb = [
            (160, 50, 40),
            (190, 100, 90),
            (180, 140, 190),
            (100, 160, 200),
            (110, 110, 110),
            (60, 150, 130),
            (50, 100, 150),
            (20, 80, 120)
        ]
       
        self.or_rectangle_images = []

        current_y_position = y0 + 120
        or_rectangle_height = 100
        or_spacing = 20
       
        for i, or_data in enumerate(data):
            if i >= 10:
                break

            color_index = i % len(or_colors_rgb)
            base_rgb = or_colors_rgb[color_index]
           
            rect_image = Image.new("RGBA", (frame_width - 60, or_rectangle_height),
                                   (base_rgb[0], base_rgb[1], base_rgb[2], 255))
            self.or_rectangle_images.append(ImageTk.PhotoImage(rect_image))
           
            rect_x0 = x0 + 30
            rect_y0 = current_y_position

            self.canvas.create_image(rect_x0, rect_y0, image=self.or_rectangle_images[-1], anchor="nw")

            self.canvas.create_text(rect_x0 + 20, rect_y0 + or_rectangle_height // 2,
                                     text=f"{or_data.get('nom', 'N/A')}",
                                     font=("Arial", 14, "bold"), anchor="w", fill="black")
           
            self.canvas.create_text(rect_x0 + 150, rect_y0 + or_rectangle_height // 2,
                                     text=f"Date limite: {or_data.get('date_limite', 'N/A')}",
                                     font=("Arial", 12), anchor="w", fill="black")

            durees = or_data.get('durees', [])
           
            durations_start_x = rect_x0 + 450
           
            for j, duree in enumerate(durees[:len(taches_display)]):
                if isinstance(duree, (int, float)):
                    formatted_duree = f"{int(duree)}h"
                else:
                    formatted_duree = f"{duree}h"

                col_offset = (j % 3) * 150
                row_offset = (j // 3) * 30
               
                self.canvas.create_text(durations_start_x + col_offset,
                                         rect_y0 + 25 + row_offset,
                                         text=f"{taches_display[j]}: {formatted_duree}",
                                         font=("Arial", 10), anchor="w", fill="black")

            current_y_position += or_rectangle_height + or_spacing

        if len(data) > 10:
            self.canvas.create_text(self.screen_width // 2, current_y_position + 20,
                                     text=f"... et {len(data)-10} ORs supplémentaires",
                                     font=("Arial", 12), fill="blue")

        btn_style = {"font": ("Arial", 12, "bold"), "width": 20, "fg": "white"}
       
        btn_retour = tk.Button(self.root, text="Retour", command=self.page2,
                                     bg="#6C757D", **btn_style)
        self.canvas.create_window(x0 + 130, y0 + frame_height - 50, window=btn_retour)

        btn_suivant_page5 = tk.Button(self.root, text="Suivant", command=lambda: self.page5(data),
                                     bg="#007BFF", **btn_style)
        self.canvas.create_window(x0 + frame_width / 2, y0 + frame_height - 50, window=btn_suivant_page5)

        btn_quitter = tk.Button(self.root, text="Quitter", command=self.quitter,
                                     bg="#DC3545", **btn_style)
        self.canvas.create_window(x0 + frame_width - 130, y0 + frame_height - 50, window=btn_quitter)


    def page5(self, donnees_ors):
        self.canvas.delete("all")

        # Fond page 5
        try:
            bg_image = Image.open("ordo.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg5'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg5'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg5: {e}")
            self.canvas.configure(bg="lightgray")

        # Cadre blanc transparent au milieu
        frame_width = 800
        frame_height = 500
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.7)))
            self.images['overlay5'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay5'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        # Titre / Question
        self.canvas.create_text(self.screen_width // 2, y0 + 100,
                                 text="Choisissez la méthode de résolution :",
                                 font=("Georgia", 24, "bold italic"), fill="black")

        # Bouton "Par heuristique (construction et amélioration)"
        btn_heuristique = tk.Button(self.root,
                                     text="Par heuristique (construction et amélioration)",
                                     command=lambda: self.trigger_heuristique_and_go_to_page6(donnees_ors),
                                     bg="#28a745", fg="white", font=("Arial", 14, "bold"), width=40)
        self.canvas.create_window(self.screen_width // 2, y0 + 200, window=btn_heuristique)

        # Bouton "Par algorithme génétique"
        btn_genetique = tk.Button(self.root,
            text="Par algorithme génétique",
            command=lambda: self.trigger_genetique_and_go_to_page6(donnees_ors),
            bg="#17a2b8", fg="white", font=("Arial", 14, "bold"), width=40)
        self.canvas.create_window(self.screen_width // 2, y0 + 280, window=btn_genetique)

        # Bouton Retour
        btn_retour_page4 = tk.Button(self.root, text="Retour", command=lambda: self.page4(donnees_ors),
                                     bg="#6C757D", fg="white", font=("Arial", 12, "bold"))
        self.canvas.create_window(x0 + 100, y0 + frame_height - 50, window=btn_retour_page4)

        # Bouton Quitter
        quit_btn = tk.Button(self.root, text="Quitter", command=self.quitter,
                               bg="#DC3545", fg="white", font=("Arial", 12, "bold"))
        self.canvas.create_window(x0 + frame_width - 100, y0 + frame_height - 50, window=quit_btn)
   
    def trigger_heuristique_and_go_to_page6(self, donnees_ors):
        # Pass input_ors_data directly to the heuristic
        scheduling_results_formatted, or_completion_times, final_machines_state, makespan = self.run_construction_heuristic(donnees_ors)
        # Marquer que c'est l'heuristique
        self.is_genetic_result = False
        # Pass all relevant data to page6
        self.page6(scheduling_results_formatted, or_completion_times, final_machines_state, donnees_ors, makespan)
        

    def trigger_genetique_and_go_to_page6(self, donnees_ors):
        try:
            print("Algorithme génétique lancé !")
            # 1. Conversion des données de l'interface au format attendu par l'algorithme génétique
            taches_template = [
                {"nom": "Demontage", "machine": None},
                {"nom": "Usinage", "machine": "Usinage"},
                {"nom": "Culasse", "machine": "Culasse"},
                {"nom": "Injection", "machine": "Injection"},
                {"nom": "Essai", "machine": None},
                {"nom": "Peinture", "machine": "Peinture"},
                {"nom": "Confection", "machine": None},
                {"nom": "Montage", "machine": None},
                {"nom": "Controle", "machine": "BEM"}
            ]
            processed_ors = []
            for i, or_data in enumerate(donnees_ors):
                try:
                    from datetime import datetime
                    date_limite_dt = datetime.strptime(or_data["date_limite"], "%d/%m/%Y")
                    delai = date_limite_dt.toordinal()
                except Exception:
                    delai = float('inf')
                taches = []
                for j, tache_template in enumerate(taches_template):
                    duree = or_data["durees"][j] if j < len(or_data["durees"]) else 0.0
                    taches.append({
                        "nom": tache_template["nom"],
                        "duree": duree,
                        "machine": tache_template["machine"]
                    })
                processed_ors.append({
                    "nom": or_data.get("nom", f"OR{i+1}"),
                    "delai": delai,
                    "poids": 1,
                    "taches": taches
                })

            # 2. Définir le template des machines
            machines_template = {
                "Usinage": [{"nom": "M1-Usinage", "libre": 0}, {"nom": "M2-Usinage", "libre": 0}],
                "Culasse": [{"nom": "M1-Culasse", "libre": 0}, {"nom": "M2-Culasse", "libre": 0}],
                "Injection": [{"nom": "M1-Inj", "libre": 0}, {"nom": "M2-Inj", "libre": 0}],
                "BEM": [{"nom": "M1-BEM", "libre": 0}, {"nom": "M2-BEM", "libre": 0}],
                "Peinture": [{"nom": "M1-Peinture", "libre": 0}, {"nom": "M2-Peinture", "libre": 0}]
            }

            # 3. Lancer l'algorithme génétique
            ga = GeneticAlgorithmORs(
                ORs_data=processed_ors,
                machines_template=machines_template,
                alpha=0.5,
                pop_size=30,
                generations=100,
                elitism_rate=0.1,
                crossover_prob=0.8,
                mutation_prob=0.15,
                max_k_mutations=2,
                max_stagnation_generations=30,
                epsilon_stagnation=0.01
            )
            meilleure_sequence_ors, meilleur_makespan = ga.run()

            # 4. Générer l'ordonnancement détaillé pour affichage
            _, ordonnancement_detail = evaluer_sequence(meilleure_sequence_ors, machines_template)

            # 5. Formater les résultats pour affichage dans page6
            scheduling_results_formatted = []
            for item in ordonnancement_detail:
                machine_info = f" ({item['Machine']})" if item["Machine"] else ""
                scheduling_results_formatted.append(
                    f"{item['OR']}-{item['Tache']}{machine_info}: {int(item['Debut'])}->{int(item['Fin'])}"
                )
            or_completion_times = {item['OR']: item['Fin'] for item in ordonnancement_detail if item['Tache'] == 'Controle'}
            final_machines_state = None
            
            # Marquer que c'est l'algorithme génétique
            self.is_genetic_result = True
            # Aller à page6
            self.page6(scheduling_results_formatted, or_completion_times, final_machines_state, donnees_ors, meilleur_makespan)
            
        except Exception as e:
            import traceback
            messagebox.showerror("Erreur", f"Erreur dans l'algorithme génétique :\n{e}\n\n{traceback.format_exc()}")


    def run_construction_heuristic(self, input_ors_data):
        # Initialisation des machines
        machines = {
            "Usinage": [{"nom": "M1-Usinage", "libre": 0}, {"nom": "M2-Usinage", "libre": 0}],
            "Culasse": [{"nom": "M1-Culasse", "libre": 0}, {"nom": "M2-Culasse", "libre": 0}],
            "Injection": [{"nom": "M1-Inj", "libre": 0}, {"nom": "M2-Inj", "libre": 0}],
            "BEM": [{"nom": "M1-BEM", "libre": 0}, {"nom": "M2-BEM", "libre": 0}],
            "Peinture": [{"nom": "M1-Peinture", "libre": 0}, {"nom": "M2-Peinture", "libre": 0}]
        }

        # Corrected template to match the input order and task names
        taches_heuristic_template_corrected = [
            {"nom": "Demontage", "machine": None},
            {"nom": "Usinage", "machine": "Usinage"},
            {"nom": "Preparation", "machine": None},
            {"nom": "Culasse", "machine": "Culasse"},
            {"nom": "Injection", "machine": "Injection"},
            {"nom": "Sous-Organe", "machine": None},
            {"nom": "Montage", "machine": None},
            {"nom": "BEM", "machine": "BEM"},
            {"nom": "Peinture", "machine": "Peinture"},
        ]
        # Convert Tkinter OR data to heuristic's OR format
        processed_ors = []
        for i, or_data_tk in enumerate(input_ors_data):
            try:
                date_limite_dt = datetime.strptime(or_data_tk["date_limite"], "%d/%m/%Y")
                delai = date_limite_dt.toordinal() # Convert to ordinal for numerical comparison
            except ValueError:
                delai = float('inf') # Assign infinite delay for invalid dates
                print(f"Warning: Invalid date format for OR {or_data_tk.get('nom', i+1)} ('{or_data_tk['date_limite']}'). Assigning infinite delay.")

            current_or_taches = []
            tk_durees = or_data_tk["durees"]
            for j, task_template in enumerate(taches_heuristic_template_corrected):
                duree = tk_durees[j] if j < len(tk_durees) else 0.0 # Ensure duree exists
                current_or_taches.append({
                    "nom": task_template["nom"],
                    "duree": duree,
                    "machine": task_template["machine"]
                })
           
            processed_ors.append({
                "nom": or_data_tk.get("nom", f"OR {i+1}"), # Use existing OR name or generate
                "delai": delai,
                "taches": current_or_taches
            })

        # Sort ORs by Earliest Due Date
        ORs_tries = sorted(processed_ors, key=lambda x: x["delai"])

        def selection_machine(machine_type, temps_dispo_or_task):
            machines_dispo = machines[machine_type]
            # Select the machine that finishes earliest, considering the OR's current earliest start time
            # A machine can only start a new task when it's free AND the OR is ready for that task.
            machine_choisie = min(machines_dispo, key=lambda m: max(m["libre"], temps_dispo_or_task))
            return machine_choisie

        ordo_detailed = []
        # Keep track of completion times for manual tasks that depend on previous manual tasks across ORs
        # For manual tasks, we assume one "worker" per task type
        temps_fin_manuelles = {
            "Demontage": 0,
            "Essai": 0,
            "Confection": 0,
            "Montage": 0
        }

        or_completion_times = {}
       
        # Max machine free time (for makespan)
        max_machine_finish_time = 0

        for OR in ORs_tries:
            # The earliest this *specific* OR can start its next task.
            # Initially, 0 for the first task, then updated after each task of this OR.
            current_or_earliest_start = 0

            for tache in OR["taches"]:
                if tache["duree"] == 0:
                    continue # Skip tasks with zero duration

                # For dependencies: A task can only start when the OR is ready AND the resource (machine or manual 'worker') is free.
                # The OR's readiness is 'current_or_earliest_start'.

                if tache["machine"] is None: # Manual task (no specific machine)
                    # This manual task can start when its "type" (e.g., Demontage) worker is free
                    # AND the OR has completed its previous task.
                    t_debut = max(temps_fin_manuelles.get(tache["nom"], 0), current_or_earliest_start)
                    t_fin = t_debut + tache["duree"]
                   
                    ordo_detailed.append({
                        "OR": OR["nom"],
                        "Tache": tache["nom"],
                        "Debut": t_debut,
                        "Fin": t_fin,
                        "Machine": "Manuel" # Indicate it's a manual process
                    })
                   
                    temps_fin_manuelles[tache["nom"]] = t_fin # Update completion time for this manual task type
                    current_or_earliest_start = t_fin # Next task in this OR can start after this manual task
                else: # Machine-dependent task
                    machine_type = tache["machine"]
                    if machine_type not in machines:
                        print(f"Warning: Machine type '{machine_type}' for task '{tache['nom']}' not defined in machines. Skipping task.")
                        continue
                   
                    # Select the best machine for this task, considering when the OR is ready (current_or_earliest_start)
                    selected_machine = selection_machine(machine_type, current_or_earliest_start)
                   
                    # Task can start when the selected machine is free AND the OR is ready
                    t_debut = max(selected_machine["libre"], current_or_earliest_start)
                    t_fin = t_debut + tache["duree"]
                   
                    selected_machine["libre"] = t_fin # Update the machine's free time
                   
                    ordo_detailed.append({
                        "OR": OR["nom"],
                        "Tache": tache["nom"],
                        "Debut": t_debut,
                        "Fin": t_fin,
                        "Machine": selected_machine["nom"]
                    })
                    current_or_earliest_start = t_fin # Next task in this OR can start after this machine task
           
            or_completion_times[OR['nom']] = current_or_earliest_start # The OR finishes when its last task finishes

        # Calculate makespan from all task end times or OR completion times
        makespan = max(t['Fin'] for t in ordo_detailed) if ordo_detailed else 0

        # Format ordo_detailed into a list of strings for display
        # This format is essential for the Treeview parsing in page6
        scheduling_results_formatted = []
        for item in ordo_detailed:
            machine_info = f" ({item['Machine']})" if item["Machine"] else ""
            scheduling_results_formatted.append(
                f"{item['OR']}-{item['Tache']}{machine_info}: {int(item['Debut'])}->{int(item['Fin'])}"
            )

        # Final state of machines (useful for debugging/analysis)
        # Deep copy to ensure the state isn't modified later by other parts of the app
        final_machines_state = copy.deepcopy(machines)
           
        return scheduling_results_formatted, or_completion_times, final_machines_state, makespan

    def page6(self, scheduling_results, or_completion_times, final_machines_state, donnees_ors_for_back, makespan_value):
        self.canvas.delete("all")

        # Fond page 6
        try:
            bg_image = Image.open("ordonn.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg6'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg6'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg6: {e}")
            self.canvas.configure(bg="lightgray")

        # Cadre blanc transparent au milieu
        frame_width = 1000
        frame_height = 700
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.7)))
            self.images['overlay6'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay6'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        # Titre
        self.canvas.create_text(self.screen_width // 2, y0 + 25,
                                 text="RÉSULTATS DE L'ORDONNANCEMENT \n PAR HEURISTIQUE DE CONSTRUCTION",
                                 font=("Georgia", 20, "bold italic"), fill="black", justify="center")

        current_y_position = y0 + 100 # Adjusted starting y for content below title

        # --- Affichage de l'ordre des ORs et du makespan ---
        ordered_ors = []
        seen_ors = set()
        for item in scheduling_results: # Use the raw string format from heuristic for parsing
            try:
                # Extract OR name from the formatted string
                # Example: "OR1-Demontage (Manuel): 0->10"
                # We need "OR1"
                or_part = item.split('-')[0]
                or_name = or_part.strip()
               
                if or_name and or_name not in seen_ors:
                    ordered_ors.append(or_name)
                    seen_ors.add(or_name)
            except IndexError:
                continue

        or_sequence_text = "Ordre des ORs : " + " -> ".join(ordered_ors) if ordered_ors else "Aucun OR ordonnancé."
        makespan_text = f"MakeSpan = {int(makespan_value)} heures" # Use makespan_value directly
       
        # Affichage de la séquence et du makespan
        self.canvas.create_text(self.screen_width // 2, current_y_position,
                                 text=or_sequence_text,
                                 font=("Arial", 14, "bold"), fill="#4B0082", anchor="center")
        current_y_position += 30
       
        self.canvas.create_text(self.screen_width // 2, current_y_position,
                                 text=makespan_text,
                                 font=("Arial", 14, "bold"), fill="#8B0000", anchor="center")
        current_y_position += 40 # Increased spacing to accommodate the Treeview title

        # Section 1: Planning détaillé des tâches
        self.canvas.create_text(self.screen_width // 2, current_y_position,
                                 text="Planning détaillé des tâches :",
                                 font=("Arial", 16, "bold"), fill="darkgreen")
        current_y_position += 10 # Space for the Treeview

        # Frame for Treeview and Scrollbar
        tree_frame_height = 400
        tree_frame_width = 700
       
        # Position the tree_frame below the text elements
        tree_frame_y = current_y_position + tree_frame_height / 2

        # IMPORTANT: Use ttk.Frame and place it on the canvas using create_window
        # and then pack the Treeview inside this frame.
        self.tree_frame_widget = ttk.Frame(self.root, width=tree_frame_width, height=tree_frame_height)
        self.canvas.create_window(self.screen_width // 2, tree_frame_y,
                                   window=self.tree_frame_widget, width=tree_frame_width, height=tree_frame_height)

        # Treeview
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=('Arial', 11, 'bold'))
        style.map('Treeview', background=[('selected', 'blue')], foreground=[('selected', 'white')])

        self.tree = ttk.Treeview(self.tree_frame_widget, columns=("OR-Tâche", "Machine", "Début", "Fin"), show="headings")
        self.tree.heading("OR-Tâche", text="OR-Tâche")
        self.tree.heading("Machine", text="Machine")
        self.tree.heading("Début", text="Début")
        self.tree.heading("Fin", text="Fin")

        self.tree.column("OR-Tâche", width=250, anchor="w")
        self.tree.column("Machine", width=150, anchor="w")
        self.tree.column("Début", width=100, anchor="center")
        self.tree.column("Fin", width=100, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.tree_frame_widget, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Define colors for ORs for Treeview rows
        or_colors_for_treeview = [
            "#FFC0CB",  # Light Pink
            "#90EE90",  # Light Green
            "#ADD8E6",  # Light Blue
            "#FFFF99",  # Light Yellow
            "#FFDAB9",  # Peach Puff
            "#D8BFD8",  # Thistle
            "#B0E0E6",  # Powder Blue
            "#F0E68C",  # Khaki
            "#FFB6C1",  # Light Pink (again, to cycle)
            "#A2D9CE"   # Medium Aquamarine
        ]
       
        # Populate Treeview
        current_or_name = None
        or_name_to_color_index = {}
        next_available_color_index = 0

        # Pre-configure all tags for colors once, before the loop
        # This is more robust than checking `style.element_names()`
        for idx, color in enumerate(or_colors_for_treeview):
            tag_name = f'or_color_{idx}'
            self.tree.tag_configure(tag_name, background=color, foreground="black") # Corrected line

        self.tree.tag_configure('separator', background='lightgray', foreground='lightgray', font=('Arial', 1, 'bold')) # Corrected line
       
        for i, line in enumerate(scheduling_results):
            parts = line.split(':')
            if len(parts) < 2:
                continue
           
            task_info = parts[0].strip()
            times_info = parts[1].strip()

            # Extract OR-Tâche and Machine
            or_task_name_full = task_info
            machine_name = "N/A" # Default for safety

            if '(' in task_info and ')' in task_info:
                or_task_name_display = task_info.split('(')[0].strip()
                machine_name = task_info.split('(')[1].replace(')', '').strip()
            else:
                or_task_name_display = task_info # No machine info in string, it's manual
                machine_name = "Manuel" # Explicitly label manual tasks

            or_name_only = or_task_name_display.split('-')[0].strip() # Extract OR name from "ORx-TaskName"
           
            start_time = "N/A"
            end_time = "N/A"
            if '->' in times_info:
                start_str, end_str = times_info.split('->')
                try:
                    start_time = int(float(start_str.strip()))
                    end_time = int(float(end_str.strip()))
                except ValueError:
                    pass

            if or_name_only not in or_name_to_color_index:
                or_name_to_color_index[or_name_only] = next_available_color_index % len(or_colors_for_treeview)
                next_available_color_index += 1
           
            color_idx = or_name_to_color_index[or_name_only]
            tag_name = f'or_color_{color_idx}'
           
            # Insert separator if new OR is encountered and it's not the very first OR
            if or_name_only != current_or_name and current_or_name is not None:
                self.tree.insert("", "end", values=("", "", "", ""), tags=('separator'))
            current_or_name = or_name_only
           
            self.tree.insert("", "end", values=(or_task_name_display, machine_name, start_time, end_time),
                                  tags=(tag_name,))
        # Update current_y_position for buttons below the Treeview
        current_y_position = tree_frame_y + tree_frame_height / 2 + 50

        # Boutons - VERSION CORRIGÉE
        btn_style_common = {"font": ("Arial", 12, "bold"), "width": 20, "fg": "white"}

        btn_retour_page5 = tk.Button(self.root, text="Retour",
                                    command=lambda: self.page5(donnees_ors_for_back),
                                    bg="#6C757D", **btn_style_common)
        self.canvas.create_window(x0 + frame_width / 2 - 250, current_y_position, window=btn_retour_page5)

        # Bouton améliorer (seulement pour l'heuristique)
        if hasattr(self, 'is_genetic_result') and not self.is_genetic_result:
            btn_ameliorer = tk.Button(self.root, text="Améliorer la solution",
                                    command=lambda: self.page7(donnees_ors_for_back, makespan_value),
                                    bg="#28A745", **btn_style_common)
            self.canvas.create_window(x0 + frame_width / 2 - 50, current_y_position, window=btn_ameliorer)

        # Bouton Gantt (pour tous les résultats)
        btn_gantt = tk.Button(self.root, text="Diagramme de Gantt",
                            command=lambda: self.afficher_gantt_simple(scheduling_results, makespan_value),
                            bg="#FF6B35", **btn_style_common)
        gantt_x_pos = x0 + frame_width / 2 + 100 if (hasattr(self, 'is_genetic_result') and not self.is_genetic_result) else x0 + frame_width / 2
        self.canvas.create_window(gantt_x_pos, current_y_position, window=btn_gantt)

        btn_quitter = tk.Button(self.root, text="Quitter", command=self.quitter,
                            bg="#DC3545", **btn_style_common)
        self.canvas.create_window(x0 + frame_width / 2 + 250, current_y_position, window=btn_quitter)

    def page7(self, donnees_ors_original, makespan_initial):
        """Page 7 - Amélioration des solutions d'ordonnancement"""
        self.canvas.delete("all")

        # Fond page 7
        try:
            bg_image = Image.open("ordonn.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg7'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg7'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg7: {e}")
            self.canvas.configure(bg="lightblue")

        # Cadre principal
        frame_width = 1000
        frame_height = 700
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2
       
        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.8)))
            self.images['overlay7'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay7'], anchor="nw")
        except:
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height, fill="white", outline="black")

        # Titre
        self.canvas.create_text(self.screen_width // 2, y0 + 30,
                               text="AMÉLIORATION DE L'ORDONNANCEMENT",
                               font=("Georgia", 22, "bold italic"), fill="darkblue", justify="center")
       
        self.canvas.create_text(self.screen_width // 2, y0 + 60,
                               text="Recherche de solutions optimisées",
                               font=("Arial", 14, "italic"), fill="darkgreen", justify="center")

        current_y = y0 + 100

        # Conversion des données pour l'algorithme d'amélioration
        processed_ors = self.convert_donnees_to_algorithm_format(donnees_ors_original)
       
        # Calcul des améliorations
        resultats_amelioration = self.calculer_ameliorations(processed_ors, makespan_initial)
       
        # Affichage des résultats
        self.afficher_resultats_amelioration(resultats_amelioration, x0, current_y, frame_width, donnees_ors_original)

     # 3. Boutons (créés après les éléments de fond)
        btn_style_common = {"font": ("Arial", 12, "bold"), "width": 20, "fg": "white"}
        button_y_position = y0 + frame_height - 50

        btn_retour_page6 = tk.Button(self.root, text="Retour",
                                        command=lambda: self.retour_page6_from_page7(donnees_ors_original),
                                        bg="#6C757D", **btn_style_common)
        self.btn_retour_page6_widget_p7 = self.canvas.create_window(x0 + frame_width / 2 - 150, button_y_position, window=btn_retour_page6)

        btn_quitter_page7 = tk.Button(self.root, text="Quitter", command=self.quitter,
                                        bg="#DC3545", **btn_style_common)
        self.btn_quitter_page7_widget_p7 = self.canvas.create_window(x0 + frame_width / 2 + 150, button_y_position, window=btn_quitter_page7)


    def afficher_gantt_simple(self, scheduling_results, makespan_value):
        """Affiche le diagramme de Gantt dans une nouvelle fenêtre"""
        try:
            # Créer une nouvelle fenêtre
            gantt_window = tk.Toplevel(self.root)
            gantt_window.title("Diagramme de Gantt")
            gantt_window.geometry("1000x700")
            
            # Importer matplotlib
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import random
            
            # Convertir les données pour le graphique
            donnees_gantt = []
            for line in scheduling_results:
                # Exemple de line: "OR1-Demontage (Manuel): 0->2"
                if ':' not in line:
                    continue
                    
                partie_gauche = line.split(':')[0]  # "OR1-Demontage (Manuel)"
                partie_droite = line.split(':')[1]  # " 0->2"
                
                # Extraire OR et Tâche
                if '-' in partie_gauche:
                    or_name = partie_gauche.split('-')[0]  # "OR1"
                    reste = partie_gauche.split('-')[1]    # "Demontage (Manuel)"
                    
                    if '(' in reste:
                        task_name = reste.split('(')[0].strip()  # "Demontage"
                        machine = reste.split('(')[1].replace(')', '').strip()  # "Manuel"
                    else:
                        task_name = reste.strip()
                        machine = "Manuel"
                else:
                    continue
                
                # Extraire début et fin
                if '->' in partie_droite:
                    debut_str = partie_droite.split('->')[0].strip()
                    fin_str = partie_droite.split('->')[1].strip()
                    try:
                        debut = int(float(debut_str))
                        fin = int(float(fin_str))
                    except:
                        continue
                else:
                    continue
                
                donnees_gantt.append({
                    'OR': or_name,
                    'Tache': task_name,
                    'Machine': machine,
                    'Debut': debut,
                    'Fin': fin
                })
            
            # Créer le graphique
            fig, ax = plt.subplots(figsize=(12, 8))
            
            # Obtenir la liste des ORs uniques
            ors_uniques = []
            for item in donnees_gantt:
                if item['OR'] not in ors_uniques:
                    ors_uniques.append(item['OR'])
            
            # Couleurs pour les tâches
            taches_uniques = []
            for item in donnees_gantt:
                if item['Tache'] not in taches_uniques:
                    taches_uniques.append(item['Tache'])
            
            couleurs = {}
            random.seed(42)
            for tache in taches_uniques:
                couleurs[tache] = (random.random(), random.random(), random.random())
            
            # Dessiner les barres
            for item in donnees_gantt:
                y_position = ors_uniques.index(item['OR'])
                duree = item['Fin'] - item['Debut']
                couleur = couleurs[item['Tache']]
                
                # Dessiner la barre
                ax.barh(y_position, duree, left=item['Debut'], height=0.5, 
                    color=couleur, edgecolor='black')
                
                # Ajouter le texte
                ax.text(item['Debut'] + duree/2, y_position, 
                    f"{item['Tache']}", 
                    ha='center', va='center', fontsize=9, weight='bold')
            
            # Configuration du graphique
            ax.set_yticks(range(len(ors_uniques)))
            ax.set_yticklabels(ors_uniques)
            ax.set_xlabel('Temps (heures)')
            ax.set_ylabel('Ordres de Réparation')
            
            # Titre différent selon la méthode
            method_name = "Algorithme Génétique" if hasattr(self, 'is_genetic_result') and self.is_genetic_result else "Heuristique de Construction"
            ax.set_title(f'Diagramme de Gantt - {method_name}\nMakespan: {int(makespan_value)} heures')
            ax.grid(axis='x', alpha=0.3)
            
            # Intégrer dans Tkinter
            canvas = FigureCanvasTkAgg(fig, gantt_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Bouton fermer
            btn_fermer = tk.Button(gantt_window, text="Fermer", 
                                command=gantt_window.destroy,
                                bg="#DC3545", fg="white", font=("Arial", 12))
            btn_fermer.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur affichage Gantt: {e}")

    def tracer_diagramme_gantt(self, ordonnancement_detail):
        """Trace le diagramme de Gantt horizontal avec la nouvelle API Matplotlib"""
        ORs_unique = sorted(list(set(t['OR'] for t in ordonnancement_detail)))
        taches_unique = sorted(list(set(t['Tache'] for t in ordonnancement_detail)))
    
        or_to_y = {or_name: i for i, or_name in enumerate(ORs_unique)}
    
        # Correction de l'avertissement de dépréciation
        colormap = plt.colormaps.get_cmap('tab20')  # Nouvelle méthode recommandée
        tache_to_color = {tache: colormap(i/len(taches_unique)) for i, tache in enumerate(taches_unique)}
    
        fig = Figure(figsize=(10, 6), dpi=100)  # Maintenant Figure est défini
        ax = fig.add_subplot(111)
    
        for tache in ordonnancement_detail:
            y_pos = or_to_y[tache['OR']]
            debut = tache['Debut']
            duree = tache['Fin'] - tache['Debut']
        
            ax.barh(y=y_pos, width=duree, left=debut,
                color=tache_to_color[tache['Tache']],
                height=0.6, edgecolor='black')
        
            if duree > 1:  # Seulement si la durée est suffisante
                ax.text(debut + duree/2, y_pos, tache['Tache'],
                    va='center', ha='center', fontsize=8)

        ax.set_yticks(list(or_to_y.values()))
        ax.set_yticklabels(list(or_to_y.keys()))
        ax.set_xlabel('Temps (heures)')
        ax.set_title('Diagramme de Gantt - Ordonnancement')
        ax.grid(axis='x', linestyle='--', alpha=0.6)
    
        return fig

    def page8(self, donnees_ors_original, makespan_initial):
        """Page 8 - Diagramme de Gantt basé sur la solution optimisée"""
        self.canvas.delete("all")

        # Fond page 8
        try:
            bg_image = Image.open("ordonn.jpg").resize((self.screen_width, self.screen_height), Image.LANCZOS)
            self.images['bg8'] = ImageTk.PhotoImage(bg_image)
            self.canvas.create_image(0, 0, image=self.images['bg8'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement bg8: {e}. Using lightblue background.")
            self.canvas.configure(bg="lightblue")

        # Cadre principal
        frame_width = 1000
        frame_height = 700
        x0 = (self.screen_width - frame_width) // 2
        y0 = (self.screen_height - frame_height) // 2

        try:
            overlay = Image.new("RGBA", (frame_width, frame_height), (255, 255, 255, int(255 * 0.8)))
            self.images['overlay8'] = ImageTk.PhotoImage(overlay)
            self.canvas.create_image(x0, y0, image=self.images['overlay8'], anchor="nw")
        except Exception as e:
            print(f"Erreur chargement overlay: {e}. Using rectangle for overlay.")
            self.canvas.create_rectangle(x0, y0, x0+frame_width, y0+frame_height,
                                        fill="white", outline="black")

        # Titre
        self.canvas.create_text(self.screen_width // 2, y0 + 30,
                                text="DIAGRAMME DE GANTT - SOLUTION OPTIMISÉE",
                                font=("Georgia", 20, "bold italic"), fill="darkblue")

        self.canvas.create_text(self.screen_width // 2, y0 + 60,
                                text="Visualisation du planning des opérations",
                                font=("Arial", 12, "italic"), fill="darkgreen")

        try:
            processed_ors = self.convert_donnees_to_algorithm_format(donnees_ors_original)
            resultats = self.calculer_ameliorations(processed_ors, makespan_initial)
        
            if resultats.get('meilleure_solution'):
                meilleure_sequence = [or_data for or_data in processed_ors
                                    if or_data['nom'] in resultats['meilleure_solution']]
                ordonnancement = self.calculer_ordonnancement_detaille(meilleure_sequence)
            
                # Cadre pour le diagramme
                frame = ttk.Frame(self.root)
                self.canvas.create_window(self.screen_width//2, self.screen_height//2,
                                        window=frame, width=1000, height=600)
            
                # Création et affichage du diagramme
                fig = self.tracer_diagramme_gantt(ordonnancement)
                canvas = FigureCanvasTkAgg(fig, master=frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            else:
                self.canvas.create_text(self.screen_width//2, self.screen_height//2,
                                    text="Aucune solution optimisée disponible",
                                    font=("Arial", 14))
        except Exception as e:
            self.canvas.create_text(self.screen_width//2, self.screen_height//2,
                                text=f"Erreur: {str(e)}",
                                font=("Arial", 14), fill="red")
            # Boutons
        btn_style_common = {"font": ("Arial", 12, "bold"), "width": 20, "fg": "white"}
        button_y_position = y0 + frame_height - 30

        # Create buttons using tk.Button and place them using create_window
        btn_retour_page7 = tk.Button(self.root, text="Retour",
                                    command=lambda: self.page7(donnees_ors_original, makespan_initial),
                                    bg="#6C757D", **btn_style_common)
        self.canvas.create_window(x0 + frame_width / 2 - 150, button_y_position, window=btn_retour_page7)

        btn_quitter = tk.Button(self.root, text="Quitter", command=self.quitter,
                                bg="#DC3545", **btn_style_common)
        self.canvas.create_window(x0 + frame_width / 2 + 150, button_y_position, window=btn_quitter)
        
    def convert_donnees_to_algorithm_format(self, donnees_ors):
        """Convertit les données de l'interface vers le format de l'algorithme d'amélioration"""
        # Template des tâches dans l'ordre correct
        taches_template = [
            {"nom": "Demontage", "machine": None},
            {"nom": "Usinage", "machine": "Usinage"},
            {"nom": "Preparation", "machine": None},
            {"nom": "Culasse", "machine": "Culasse"},
            {"nom": "Injection", "machine": "Injection"},
            {"nom": "Sous-Organe", "machine": None},
            {"nom": "Montage", "machine": None},
            {"nom": "BEM", "machine": "BEM"},
            {"nom": "Peinture", "machine": "Peinture"},
        ]
        processed_ors = []
        for i, or_data in enumerate(donnees_ors):
            try:
                from datetime import datetime
                date_limite_dt = datetime.strptime(or_data["date_limite"], "%d/%m/%Y")
                delai = date_limite_dt.toordinal()
            except ValueError:
                delai = float('inf')
               
            taches_or = []
            durees = or_data["durees"]
           
            for j, tache_template in enumerate(taches_template):
                duree = durees[j] if j < len(durees) else 0.0
                taches_or.append({
                    "nom": tache_template["nom"],
                    "duree": duree,
                    "machine": tache_template["machine"]
                })
           
            processed_ors.append({
                "nom": or_data.get("nom", f"OR {i+1}"),
                "delai": delai,
                "taches": taches_or
            })
       
        return processed_ors

    def calculer_makespan_et_retards(self, OR_sequence):
        """Calcule le makespan et les retards pour une séquence donnée"""
        machines_template = {
            "Usinage": [{"nom": "M1-Usinage", "libre": 0}, {"nom": "M2-Usinage", "libre": 0}],
            "Culasse": [{"nom": "M1-Culasse", "libre": 0}, {"nom": "M2-Culasse", "libre": 0}],
            "Injection": [{"nom": "M1-Inj", "libre": 0}, {"nom": "M2-Inj", "libre": 0}],
            "BEM": [{"nom": "M1-BEM", "libre": 0}, {"nom": "M2-BEM", "libre": 0}],
            "Peinture": [{"nom": "M1-Peinture", "libre": 0}, {"nom": "M2-Peinture", "libre": 0}]
        }
       
        machines = copy.deepcopy(machines_template)
        temps_manuelles = {"Demontage": 0, "Essai": 0, "Confection": 0, "Montage": 0}
        makespan = 0
        retards = {}
       
        for OR in OR_sequence:
            t = 0
            for tache in OR["taches"]:
                if tache["duree"] == 0:
                    continue
                   
                if tache["machine"]:
                    machines_dispo = machines[tache["machine"]]
                    machine = min(machines_dispo, key=lambda m: max(m["libre"], t))
                    debut = max(machine["libre"], t)
                    fin = debut + tache["duree"]
                    machine["libre"] = fin
                else:
                    debut = max(temps_manuelles.get(tache["nom"], 0), t)
                    fin = debut + tache["duree"]
                    temps_manuelles[tache["nom"]] = fin
                t = fin
           
            retards[OR["nom"]] = t - OR["delai"]
            makespan = max(makespan, t)
       
        return makespan, retards

    def corriger_retards(self, sequence):
        """Algorithme de correction des retards"""
        makespan, retards = self.calculer_makespan_et_retards(sequence)
        sequence_corrigee = sequence.copy()
        amelioration = True
        iterations = 0
        max_iterations = 50  # Limite pour éviter les boucles infinies
       
        while amelioration and iterations < max_iterations:
            amelioration = False
            iterations += 1
           
            if not retards:
                break
               
            or_retarde = max(retards.items(), key=lambda x: x[1])
            or_avance = min(retards.items(), key=lambda x: x[1])
           
            if or_retarde[1] <= 0 or or_avance[1] >= 0:
                break
           
            try:
                idx_retarde = next(i for i, or_ in enumerate(sequence_corrigee) if or_["nom"] == or_retarde[0])
                idx_avance = next(i for i, or_ in enumerate(sequence_corrigee) if or_["nom"] == or_avance[0])
            except StopIteration:
                break
           
            new_seq = sequence_corrigee.copy()
            new_seq[idx_retarde], new_seq[idx_avance] = new_seq[idx_avance], new_seq[idx_retarde]
            new_makespan, new_retards = self.calculer_makespan_et_retards(new_seq)
           
            if new_makespan < makespan:
                sequence_corrigee = new_seq
                makespan = new_makespan
                retards = new_retards
                amelioration = True
       
        return sequence_corrigee, makespan

    def calculer_ameliorations(self, processed_ors, makespan_initial):
        """Calcule les améliorations possibles"""
        # Tri EDD initial
        ors_edd = sorted(processed_ors, key=lambda x: x["delai"])
        makespan_edd, _ = self.calculer_makespan_et_retards(ors_edd)
       
        # Recherche de la meilleure solution
        or_urgent = min(processed_ors, key=lambda x: x["delai"])
        autres_ors = [or_ for or_ in processed_ors if or_["nom"] != or_urgent["nom"]]
       
        meilleure_solution = None
        meilleur_makespan = makespan_edd
        solutions_testees = []
       
        # Test de différentes permutations (limité pour l'interface)
        max_permutations = min(24, len(list(itertools.permutations(autres_ors))))  # Max 24 permutations
       
        for i, perm in enumerate(itertools.permutations(autres_ors)):
            if i >= max_permutations:
                break
               
            sequence_init = [or_urgent] + list(perm)
            makespan_init, _ = self.calculer_makespan_et_retards(sequence_init)
            sequence_corr, makespan_corr = self.corriger_retards(sequence_init)
           
            solutions_testees.append({
                'sequence_init': [or_['nom'] for or_ in sequence_init],
                'makespan_init': makespan_init,
                'sequence_corr': [or_['nom'] for or_ in sequence_corr],
                'makespan_corr': makespan_corr
            })
           
            if makespan_corr < meilleur_makespan:
                meilleure_solution = sequence_corr
                meilleur_makespan = makespan_corr
       
        return {
            'edd_sequence': [or_['nom'] for or_ in ors_edd],
            'makespan_edd': makespan_edd,
            'makespan_initial': makespan_initial,
            'meilleure_solution': [or_['nom'] for or_ in meilleure_solution] if meilleure_solution else None,
            'meilleur_makespan': meilleur_makespan,
            'solutions_testees': solutions_testees,
            'amelioration_trouvee': meilleure_solution is not None and meilleur_makespan < makespan_initial
        }

    def afficher_resultats_amelioration(self, resultats, x0, start_y, frame_width, donnees_ors_original):
        """Affiche les résultats d'amélioration dans l'interface"""
        current_y = start_y
       
        # Résumé des résultats
        self.canvas.create_text(x0 + frame_width // 2, current_y,
                               text="📊 RÉSULTATS DE L'ANALYSE",
                               font=("Arial", 16, "bold"), fill="darkblue")
        current_y += 40
       
        # Makespan initial vs amélioré
        texte_makespan = f"Makespan initial (EDD): {int(resultats['makespan_edd'])} heures"
        self.canvas.create_text(x0 + 50, current_y, text=texte_makespan,
                               font=("Arial", 12), fill="black", anchor="w")
        current_y += 25
       
        if resultats['amelioration_trouvee']:
            gain = resultats['makespan_edd'] - resultats['meilleur_makespan']
            pourcentage = (gain / resultats['makespan_edd']) * 100
           
            texte_ameliore = f"✅ Meilleur makespan trouvé: {int(resultats['meilleur_makespan'])} heures"
            self.canvas.create_text(x0 + 50, current_y, text=texte_ameliore,
                                   font=("Arial", 12, "bold"), fill="darkgreen", anchor="w")
            current_y += 25
           
            texte_gain = f"🎯 Gain: {int(gain)} heures ({pourcentage:.1f}% d'amélioration)"
            self.canvas.create_text(x0 + 50, current_y, text=texte_gain,
                                   font=("Arial", 12, "bold"), fill="red", anchor="w")
            current_y += 35
           
            # Séquences
            seq_edd = " → ".join(resultats['edd_sequence'])
            seq_amelioree = " → ".join(resultats['meilleure_solution'])
           
            self.canvas.create_text(x0 + 50, current_y, text="Séquence EDD:",
                                   font=("Arial", 11, "bold"), fill="black", anchor="w")
            current_y += 15
            self.canvas.create_text(x0 + 70, current_y, text=seq_edd,
                                   font=("Arial", 10), fill="darkblue", anchor="w")
            current_y += 20
           
            self.canvas.create_text(x0 + 50, current_y, text="Séquence améliorée:",
                                   font=("Arial", 11, "bold"), fill="black", anchor="w")
            current_y += 15
            self.canvas.create_text(x0 + 70, current_y, text=seq_amelioree,
                                   font=("Arial", 10), fill="darkgreen", anchor="w")
            current_y += 35
           
            # Affichage des détails de la solution améliorée
            self.afficher_details_solution_amelioree(resultats, x0, current_y, frame_width, donnees_ors_original)
            current_y += 450  # Espace pour le tableau détaillé
           
        else:
            self.canvas.create_text(x0 + 50, current_y,
                                   text="❌ Aucune amélioration trouvée avec les permutations testées",
                                   font=("Arial", 12), fill="darkorange", anchor="w")
            current_y += 80
       
     
    def afficher_details_solution_amelioree(self, resultats, x0, start_y, frame_width, donnees_ors_original):
        """Affiche le détail de l'ordonnancement de la solution améliorée"""
        current_y = start_y
       
        # Recalcul de l'ordonnancement détaillé pour la meilleure solution
        processed_ors = self.convert_donnees_to_algorithm_format(donnees_ors_original)
       
        # Réorganiser selon la meilleure séquence
        meilleure_sequence_ors = []
        for nom_or in resultats['meilleure_solution']:
            for or_data in processed_ors:
                if or_data['nom'] == nom_or:
                    meilleure_sequence_ors.append(or_data)
                    break
       
        # Calculer l'ordonnancement détaillé
        ordo_ameliore = self.calculer_ordonnancement_detaille(meilleure_sequence_ors)
       
        # Frame pour le tableau des détails
        details_frame = ttk.Frame(self.root)
        details_height = 250
        self.canvas.create_window(x0 + frame_width // 2, current_y + details_height // 2,
                                 window=details_frame, width=frame_width - 50, height=details_height)
       
        # Treeview pour l'ordonnancement détaillé
        style = ttk.Style()
        style.configure("Details.Treeview", rowheight=25)
        style.configure("Details.Treeview.Heading", font=('Arial', 11, 'bold'))
       
        columns_details = ("OR-Tâche", "Machine", "Début", "Fin", "Durée")
        tree_details = ttk.Treeview(details_frame, columns=columns_details, show="headings",
                                   height=12, style="Details.Treeview")
       
        tree_details.heading("OR-Tâche", text="OR-Tâche")
        tree_details.heading("Machine", text="Machine")
        tree_details.heading("Début", text="Début")
        tree_details.heading("Fin", text="Fin")
        tree_details.heading("Durée", text="Durée")
       
        tree_details.column("OR-Tâche", width=200, anchor="w")
        tree_details.column("Machine", width=120, anchor="w")
        tree_details.column("Début", width=80, anchor="center")
        tree_details.column("Fin", width=80, anchor="center")
        tree_details.column("Durée", width=80, anchor="center")
       
        # Scrollbar pour les détails
        scrollbar_details = ttk.Scrollbar(details_frame, orient="vertical", command=tree_details.yview)
        tree_details.configure(yscrollcommand=scrollbar_details.set)
       
        tree_details.pack(side="left", fill="both", expand=True)
        scrollbar_details.pack(side="right", fill="y")
       
        # Couleurs pour différencier les ORs
        or_colors = [
"#FFC0CB",  # Light Pink
            "#90EE90",  # Light Green
            "#ADD8E6",  # Light Blue
            "#FFFF99",  # Light Yellow
            "#FFDAB9",  # Peach Puff
            "#D8BFD8",  # Thistle
            "#B0E0E6",  # Powder Blue
            "#F0E68C",  # Khaki
            "#FFB6C1",  # Light Pink (again, to cycle)
            "#A2D9CE"   # Medium Aquamarine
        ]

       
        # Configuration des tags de couleur
        for i, color in enumerate(or_colors):
            tree_details.tag_configure(f'or_color_{i}', background=color, foreground="black")
       
        tree_details.tag_configure('separator', background='#D3D3D3', foreground='#D3D3D3')
       
        # Remplir le tableau avec l'ordonnancement détaillé
        current_or = None
        or_color_map = {}
        color_index = 0
       
        for item in ordo_ameliore:
            or_name = item["OR"]
           
            # Assigner une couleur à chaque OR
            if or_name not in or_color_map:
                or_color_map[or_name] = color_index % len(or_colors)
                color_index += 1
           
            # Ajouter un séparateur entre les ORs
            if current_or is not None and current_or != or_name:
                tree_details.insert("", "end", values=("", "", "", "", ""), tags=('separator',))
           
            current_or = or_name
           
            # Calculer la durée
            duree = item["Fin"] - item["Debut"]
           
            # Formater le nom de la tâche
            tache_complete = f"{item['OR']}-{item['Tache']}"
           
            # Insérer la ligne avec la couleur appropriée
            color_tag = f'or_color_{or_color_map[or_name]}'
            tree_details.insert("", "end",
                               values=(tache_complete, item["Machine"],
                                      int(item["Debut"]), int(item["Fin"]), int(duree)),
                               tags=(color_tag,))

    def calculer_ordonnancement_detaille(self, OR_sequence):
        """Calcule l'ordonnancement détaillé pour une séquence donnée"""
        machines_template = {
            "Usinage": [{"nom": "M1-Usinage", "libre": 0}, {"nom": "M2-Usinage", "libre": 0}],
            "Culasse": [{"nom": "M1-Culasse", "libre": 0}, {"nom": "M2-Culasse", "libre": 0}],
            "Injection": [{"nom": "M1-Inj", "libre": 0}, {"nom": "M2-Inj", "libre": 0}],
            "BEM": [{"nom": "M1-BEM", "libre": 0}, {"nom": "M2-BEM", "libre": 0}],
            "Peinture": [{"nom": "M1-Peinture", "libre": 0}, {"nom": "M2-Peinture", "libre": 0}]
        }
       
        machines = copy.deepcopy(machines_template)
        temps_manuelles = {"Demontage": 0, "Essai": 0, "Confection": 0, "Montage": 0}
        ordo_detailed = []
       
        for OR in OR_sequence:
            current_or_time = 0
           
            for tache in OR["taches"]:
                if tache["duree"] == 0:
                    continue
                   
                if tache["machine"]:  # Tâche machine
                    machines_dispo = machines[tache["machine"]]
                    machine = min(machines_dispo, key=lambda m: max(m["libre"], current_or_time))
                   
                    debut = max(machine["libre"], current_or_time)
                    fin = debut + tache["duree"]
                    machine["libre"] = fin
                   
                    ordo_detailed.append({
                        "OR": OR["nom"],
                        "Tache": tache["nom"],
                        "Debut": debut,
                        "Fin": fin,
                        "Machine": machine["nom"]
                    })
                   
                else:  # Tâche manuelle
                    debut = max(temps_manuelles.get(tache["nom"], 0), current_or_time)
                    fin = debut + tache["duree"]
                    temps_manuelles[tache["nom"]] = fin
                   
                    ordo_detailed.append({
                        "OR": OR["nom"],
                        "Tache": tache["nom"],
                        "Debut": debut,
                        "Fin": fin,
                        "Machine": "Manuel"
                    })
               
                current_or_time = fin
       
        return ordo_detailed
        """Retour à la page 6 depuis la page 7"""
        # Recalculer les résultats pour page6
        scheduling_results, or_completion_times, final_machines_state, makespan = self.run_construction_heuristic(donnees_ors)
        self.page6(scheduling_results, or_completion_times, final_machines_state, donnees_ors, makespan)

    def appliquer_solution_amelioree(self, sequence_amelioree, donnees_ors_original):
        """Applique la solution améliorée et retourne aux résultats"""
        try:
            # Réorganiser les données selon la nouvelle séquence
            donnees_reordonnees = []
           
            for nom_or in sequence_amelioree:
                for or_data in donnees_ors_original:
                    if or_data.get("nom", "") == nom_or:
                        donnees_reordonnees.append(or_data)
                        break
           
            # Si certains ORs n'ont pas été trouvés, les ajouter à la fin
            noms_trouves = [or_data.get("nom", "") for or_data in donnees_reordonnees]
            for or_data in donnees_ors_original:
                if or_data.get("nom", "") not in noms_trouves:
                    donnees_reordonnees.append(or_data)
           
            # Recalculer avec la nouvelle séquence
            scheduling_results, or_completion_times, final_machines_state, makespan = self.run_construction_heuristic(donnees_reordonnees)
           
            # Afficher les résultats améliorés
            messagebox.showinfo("Solution appliquée",
                               f"✅ Solution améliorée appliquée avec succès!\n"
                               f"Nouveau makespan: {int(makespan)} heures")
           
            self.page6(scheduling_results, or_completion_times, final_machines_state, donnees_reordonnees, makespan)
           
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'application de la solution: {str(e)}")

   

# Lancement de l'application
if __name__ == "__main__":
    root = tk.Tk()
    app = Application(root)
    root.mainloop()